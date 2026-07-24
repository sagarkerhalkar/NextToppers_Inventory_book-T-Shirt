import calendar
from datetime import date, datetime, time, timedelta
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import get_object_or_404
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import BookAllocation, Employee, TshirtAllocation

TEAL = "0F766E"
DARK_GRAY = "263238"


def _excel_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.strftime("%d-%m-%Y %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%d-%m-%Y")
    return value


def _style_sheet(sheet):
    fill = PatternFill("solid", fgColor=TEAL)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        sheet.column_dimensions[column[0].column_letter].width = min(max(width, 11), 45)


def _append_sheet(workbook, title, headers, rows):
    sheet = workbook.create_sheet(title[:31])
    sheet.append(headers)
    for row in rows:
        sheet.append([_excel_value(value) for value in row])
    _style_sheet(sheet)


def _parse_date(value, name):
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{name} must use YYYY-MM-DD format.") from exc


def _period(period_name, start_value="", end_value="", month_value="", year_value=""):
    today = timezone.localdate()
    definitions = {
        "month": (today.replace(day=1), today, "Current month"),
        "quarter": (today - timedelta(days=89), today, "Quarterly — rolling 90 days"),
        "half_year": (today - timedelta(days=179), today, "Half-yearly — rolling 180 days"),
        "year": (today - timedelta(days=364), today, "Yearly — rolling 365 days"),
        "all": (date(2000, 1, 1), today, "All available history"),
    }
    if period_name == "custom":
        if not start_value or not end_value:
            raise ValueError("Custom report requires start date and end date.")
        start_date = _parse_date(start_value, "Start date")
        end_date = _parse_date(end_value, "End date")
        label = "Custom period"
    elif period_name == "calendar_month":
        try:
            selected = datetime.strptime(month_value, "%Y-%m").date()
        except (TypeError, ValueError) as exc:
            raise ValueError("Select a month using YYYY-MM format.") from exc
        start_date = selected.replace(day=1)
        end_date = selected.replace(day=calendar.monthrange(selected.year, selected.month)[1])
        label = selected.strftime("Calendar month — %B %Y")
    elif period_name == "calendar_year":
        try:
            selected_year = int(year_value)
        except (TypeError, ValueError) as exc:
            raise ValueError("Enter a valid four-digit year.") from exc
        if selected_year < 2000 or selected_year > today.year:
            raise ValueError(f"Year must be between 2000 and {today.year}.")
        start_date = date(selected_year, 1, 1)
        end_date = date(selected_year, 12, 31)
        label = f"Calendar year — {selected_year}"
    elif period_name in definitions:
        start_date, end_date, label = definitions[period_name]
    else:
        raise ValueError("Unknown report period.")
    if start_date > end_date:
        raise ValueError("Start date cannot be after end date.")
    if end_date > today:
        end_date = today
    tz = timezone.get_current_timezone()
    start_dt = timezone.make_aware(datetime.combine(start_date, time.min), tz)
    end_dt = timezone.make_aware(datetime.combine(end_date + timedelta(days=1), time.min), tz)
    return start_date, end_date, start_dt, end_dt, label


def _book_queryset(start_dt, end_dt, employee=None):
    queryset = (
        BookAllocation.objects.select_related(
            "book", "employee", "employee_record", "allocated_by", "returned_by"
        )
        .filter(
            Q(allocated_at__gte=start_dt, allocated_at__lt=end_dt)
            | Q(returned_at__gte=start_dt, returned_at__lt=end_dt)
        )
        .order_by("-allocated_at")
    )
    if employee is not None:
        queryset = queryset.filter(employee_record=employee)
    return queryset


def _tshirt_queryset(start_dt, end_dt, employee=None):
    queryset = (
        TshirtAllocation.objects.select_related(
            "employee", "employee_record", "stock", "stock__brand", "requested_by", "issued_by"
        )
        .filter(
            Q(requested_at__gte=start_dt, requested_at__lt=end_dt)
            | Q(issued_at__gte=start_dt, issued_at__lt=end_dt)
        )
        .order_by("-requested_at")
    )
    if employee is not None:
        queryset = queryset.filter(employee_record=employee)
    return queryset


def _book_rows(queryset):
    rows = []
    for item in queryset:
        employee = item.recipient
        rows.append((
            item.book.asset_id,
            item.book.name,
            item.book.publication_name,
            item.book.subject,
            employee.employee_id if employee else "",
            employee.full_name if employee else "",
            item.allocated_at,
            item.allocated_by.employee_id if item.allocated_by else "",
            item.returned_at,
            item.returned_by.employee_id if item.returned_by else "",
            item.get_return_condition_display() if item.return_condition else "",
            item.return_note,
            "Currently Allocated" if item.is_active else "Returned",
        ))
    return rows


def _tshirt_rows(queryset):
    rows = []
    for item in queryset:
        employee = item.recipient
        rows.append((
            employee.employee_id if employee else "",
            employee.full_name if employee else "",
            item.stock.brand.name,
            item.stock.size,
            item.quantity,
            item.get_issue_type_display(),
            item.get_status_display(),
            item.requested_at,
            item.issued_at,
            item.requested_by.employee_id if item.requested_by else "",
            item.issued_by.employee_id if item.issued_by else "",
            item.payment_amount,
        ))
    return rows


def _pdf_table(data, widths):
    table = Table(data, repeatRows=1, colWidths=widths)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#B0BEC5")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#E8F5F3")]),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 6.2),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    return table


@login_required
def download_activity_report(request):
    report_type = request.GET.get("report_type", "combined")
    period_name = request.GET.get("period", "month")
    output_format = request.GET.get("format", "xlsx")
    employee_value = request.GET.get("employee", "").strip()
    if report_type not in {"book", "tshirt", "combined"}:
        return HttpResponseBadRequest("Unknown report type.")
    if output_format not in {"xlsx", "pdf"}:
        return HttpResponseBadRequest("Unknown report format.")
    employee = None
    if employee_value:
        try:
            employee = get_object_or_404(Employee, pk=int(employee_value))
        except (TypeError, ValueError):
            return HttpResponseBadRequest("Invalid employee selection.")
    try:
        start_date, end_date, start_dt, end_dt, label = _period(
            period_name,
            request.GET.get("start_date", ""),
            request.GET.get("end_date", ""),
            request.GET.get("month", ""),
            request.GET.get("year", ""),
        )
    except ValueError as exc:
        return HttpResponseBadRequest(str(exc))

    books = _book_rows(_book_queryset(start_dt, end_dt, employee)) if report_type in {"book", "combined"} else []
    tshirts = _tshirt_rows(_tshirt_queryset(start_dt, end_dt, employee)) if report_type in {"tshirt", "combined"} else []
    slug = f"{start_date.isoformat()}_to_{end_date.isoformat()}"
    employee_label = f"{employee.employee_id} — {employee.full_name}" if employee else "All employees"
    employee_slug = employee.employee_id if employee else "all_employees"
    title = f"Next Toppers Inventory Activity — {label}"

    if output_format == "xlsx":
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"
        summary.append(["Report", title])
        summary.append(["Employee", employee_label])
        summary.append(["From", start_date.strftime("%d-%m-%Y")])
        summary.append(["To", end_date.strftime("%d-%m-%Y")])
        summary.append(["Book activity rows", len(books)])
        summary.append(["T-shirt activity rows", len(tshirts)])
        summary.append(["T-shirts in report", sum(row[4] for row in tshirts)])
        for cell in summary["A"]:
            cell.font = Font(bold=True, color=DARK_GRAY)
        summary.column_dimensions["A"].width = 24
        summary.column_dimensions["B"].width = 52
        if report_type in {"book", "combined"}:
            _append_sheet(workbook, "Book Activity", [
                "Asset ID", "Book", "Publication", "Subject", "Employee ID", "Employee", "Allocated at", "Allocated by",
                "Returned at", "Returned by", "Return condition", "Return note", "Status",
            ], books)
        if report_type in {"tshirt", "combined"}:
            _append_sheet(workbook, "T-shirt Activity", [
                "Employee ID", "Employee", "Brand", "Size", "Quantity", "Type", "Status",
                "Requested at", "Issued at", "Entered by", "Issued by", "Payment amount",
            ], tshirts)
        output = BytesIO()
        workbook.save(output)
        response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{employee_slug}_{report_type}_inventory_{slug}.xlsx"'
        return response

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{employee_slug}_{report_type}_inventory_{slug}.pdf"'
    document = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(title, styles["Title"]),
        Paragraph(f"Employee: {employee_label}", styles["Normal"]),
        Paragraph(f"Period: {start_date.strftime('%d-%m-%Y')} to {end_date.strftime('%d-%m-%Y')}", styles["Normal"]),
        Spacer(1, 7*mm),
    ]
    if report_type in {"book", "combined"}:
        story.append(Paragraph(f"Book Activity ({len(books)})", styles["Heading2"]))
        data = [["Asset ID", "Book", "Publication", "Subject", "Employee", "Allocated", "Returned", "Status"]]
        data.extend([[row[0], row[1], row[2], row[3], f"{row[4]} {row[5]}", _excel_value(row[6]), _excel_value(row[8]), row[12]] for row in books])
        story.append(_pdf_table(data, [20*mm, 43*mm, 35*mm, 28*mm, 40*mm, 33*mm, 33*mm, 26*mm]))
    if report_type == "combined":
        story.append(PageBreak())
    if report_type in {"tshirt", "combined"}:
        story.append(Paragraph(f"T-shirt Activity ({len(tshirts)})", styles["Heading2"]))
        data = [["Employee", "Brand", "Size", "Qty", "Type", "Status", "Requested", "Issued", "Entered by"]]
        data.extend([[f"{row[0]} {row[1]}", row[2], row[3], row[4], row[5], row[6], _excel_value(row[7]), _excel_value(row[8]), row[9]] for row in tshirts])
        story.append(_pdf_table(data, [43*mm, 36*mm, 16*mm, 12*mm, 20*mm, 24*mm, 36*mm, 36*mm, 25*mm]))
    document.build(story)
    return response
