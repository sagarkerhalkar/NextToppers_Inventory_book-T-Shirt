from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas

from .models import Book


@login_required
def export_books_excel(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Book Inventory"
    headers = [
        "Book Number / Asset ID",
        "Book Name",
        "Publication Name",
        "Subject",
        "Class",
        "Stream",
        "ISBN",
        "Condition",
        "Status",
    ]
    sheet.append(headers)
    for book in Book.objects.filter(is_active=True).order_by("asset_id"):
        sheet.append([
            book.asset_id,
            book.name,
            book.publication_name,
            book.subject,
            book.class_name,
            book.stream_name,
            book.isbn,
            book.get_condition_display(),
            book.get_status_display(),
        ])
    header_fill = PatternFill("solid", fgColor="0F766E")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        sheet.column_dimensions[column[0].column_letter].width = min(max(width, 12), 42)
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="book_inventory.xlsx"'
    return response


@login_required
def export_books_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="book_inventory.pdf"'
    pdf = canvas.Canvas(response, pagesize=landscape(A4))
    width, height = landscape(A4)
    y = height - 38

    def draw_header():
        nonlocal y
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(32, y, "Next Toppers Book Inventory")
        y -= 24
        pdf.setFont("Helvetica-Bold", 6.5)
        pdf.drawString(32, y, "Book No. | Book | Publication | Subject | Class/Stream | ISBN | Condition | Status")
        y -= 16
        pdf.line(32, y + 5, width - 32, y + 5)
        pdf.setFont("Helvetica", 6.5)

    draw_header()
    for book in Book.objects.filter(is_active=True).order_by("asset_id")[:10000]:
        if y < 30:
            pdf.showPage()
            y = height - 38
            draw_header()
        line = (
            f"{book.asset_id} | {book.name[:32]} | {book.publication_name[:22]} | "
            f"{book.subject[:18]} | {book.class_name}/{book.stream_name[:12]} | "
            f"{book.isbn[:18]} | {book.get_condition_display()} | {book.get_status_display()}"
        )
        pdf.drawString(32, y, line)
        y -= 13
    pdf.save()
    return response
