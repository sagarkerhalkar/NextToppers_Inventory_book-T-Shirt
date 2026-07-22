from datetime import date, datetime
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas

from .models import AuditLog, BookAllocation, Employee, TshirtAllocation, TshirtStock, User
from .permissions import role_required


def _excel_response(filename, sheet_name, headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name[:31]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in rows:
        normalised = []
        for value in row:
            if value is None:
                normalised.append("")
            elif isinstance(value, datetime):
                if timezone.is_aware(value):
                    value = timezone.localtime(value)
                normalised.append(value.strftime("%d-%m-%Y %H:%M:%S"))
            elif isinstance(value, date):
                normalised.append(value.strftime("%d-%m-%Y"))
            else:
                normalised.append(value)
        sheet.append(normalised)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        sheet.column_dimensions[column[0].column_letter].width = min(max(width, 10), 45)
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _pdf_response(filename, title, headers, rows, widths=None):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    page_size = landscape(A4)
    pdf = canvas.Canvas(response, pagesize=page_size)
    width, height = page_size
    x_start = 32
    y = height - 36
    row_height = 15
    widths = widths or [max((width - 64) / len(headers), 60)] * len(headers)

    def draw_header():
        nonlocal y
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(x_start, y, title)
        y -= 24
        pdf.setFont("Helvetica-Bold", 7)
        x = x_start
        for header, column_width in zip(headers, widths):
            pdf.drawString(x, y, str(header)[:30])
            x += column_width
        y -= row_height
        pdf.line(x_start, y + 4, width - 32, y + 4)

    draw_header()
    pdf.setFont("Helvetica", 7)
    for row in rows:
        if y < 35:
            pdf.showPage()
            y = height - 36
            draw_header()
            pdf.setFont("Helvetica", 7)
        x = x_start
        for value, column_width in zip(row, widths):
            text = str(value or "")
            max_chars = max(int(column_width / 4.2), 8)
            pdf.drawString(x, y, text[:max_chars])
            x += column_width
        y -= row_height
    pdf.save()
    return response


@login_required
def export_employees_excel(request):
    rows = Employee.objects.order_by("employee_id").values_list("employee_id", "full_name", "mobile_number", "email", "department", "designation", "joining_date", "office_location", "default_tshirt_size", "is_active", "notes")
    return _excel_response("employee_master.xlsx", "Employees", ["Employee ID", "Full name", "Mobile", "Email", "Department", "Designation", "Joining date", "Office", "T-shirt size", "Active", "Notes"], rows)


@login_required
def export_book_history_excel(request):
    rows = []
    for item in BookAllocation.objects.select_related("book", "employee", "employee_record", "allocated_by", "returned_by"):
        employee = item.recipient
        rows.append((item.book.asset_id, item.book.name, employee.employee_id if employee else "", employee.full_name if employee else "", item.allocated_at, item.allocated_by.employee_id if item.allocated_by else "", item.returned_at, item.return_condition, item.return_note, item.returned_by.employee_id if item.returned_by else "", item.is_active))
    return _excel_response("book_allocation_return_history.xlsx", "Book History", ["Asset ID", "Book", "Employee ID", "Employee", "Allocated at", "Allocated by", "Returned at", "Return condition", "Return note", "Returned by", "Active allocation"], rows)


@login_required
def export_book_history_pdf(request):
    rows = []
    for item in BookAllocation.objects.select_related("book", "employee", "employee_record"):
        employee = item.recipient
        rows.append((item.book.asset_id, item.book.name, employee.employee_id if employee else "", timezone.localtime(item.allocated_at).strftime("%d-%m-%Y %H:%M") if item.allocated_at else "", timezone.localtime(item.returned_at).strftime("%d-%m-%Y %H:%M") if item.returned_at else "", item.return_condition, "Open" if item.is_active else "Returned"))
    return _pdf_response("book_allocation_return_history.pdf", "Next Toppers Book Allocation and Return History", ["Asset ID", "Book", "Employee", "Allocated", "Returned", "Condition", "Status"], rows, [70, 190, 80, 105, 105, 65, 55])


@login_required
def export_tshirt_allocations_excel(request):
    rows = []
    for item in TshirtAllocation.objects.select_related("employee", "employee_record", "stock", "stock__brand", "requested_by", "approved_by", "issued_by"):
        employee = item.recipient
        rows.append((employee.employee_id if employee else "", employee.full_name if employee else "", item.stock.brand.name, item.stock.size, item.quantity, item.issue_type, item.status, item.requested_at, item.requested_by.employee_id if item.requested_by else "", item.payment_amount, item.payment_date, item.approved_by.employee_id if item.approved_by else "", item.approved_at, item.issued_by.employee_id if item.issued_by else "", item.issued_at, item.rejection_reason))
    return _excel_response("tshirt_allocation_history.xlsx", "T-shirt History", ["Employee ID", "Employee", "Brand", "Size", "Quantity", "Issue type", "Status", "Requested at", "Requested by", "Payment amount", "Payment date", "Approved by", "Approved at", "Issued by", "Issued at", "Rejection reason"], rows)


@login_required
def export_employee_history_excel(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    workbook = Workbook()
    book_sheet = workbook.active
    book_sheet.title = "Book History"
    book_sheet.append(["Asset ID", "Book", "Allocated at", "Allocated by", "Returned at", "Returned by", "Condition", "Return note", "Status"])
    for item in employee.book_allocations.select_related("book", "allocated_by", "returned_by"):
        book_sheet.append([item.book.asset_id, item.book.name, timezone.localtime(item.allocated_at).strftime("%d-%m-%Y %H:%M:%S") if item.allocated_at else "", item.allocated_by.employee_id if item.allocated_by else "", timezone.localtime(item.returned_at).strftime("%d-%m-%Y %H:%M:%S") if item.returned_at else "", item.returned_by.employee_id if item.returned_by else "", item.return_condition, item.return_note, "Currently Allocated" if item.is_active else "Returned"])
    tshirt_sheet = workbook.create_sheet("T-shirt History")
    tshirt_sheet.append(["Brand", "Size", "Quantity", "Type", "Status", "Requested at", "Issued at", "Issued by"])
    for item in employee.tshirt_allocations.select_related("stock", "stock__brand", "issued_by"):
        tshirt_sheet.append([item.stock.brand.name, item.stock.size, item.quantity, item.issue_type, item.status, timezone.localtime(item.requested_at).strftime("%d-%m-%Y %H:%M:%S") if item.requested_at else "", timezone.localtime(item.issued_at).strftime("%d-%m-%Y %H:%M:%S") if item.issued_at else "", item.issued_by.employee_id if item.issued_by else ""])
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column) + 2
            sheet.column_dimensions[column[0].column_letter].width = min(max(width, 10), 42)
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{employee.employee_id}_inventory_history.xlsx"'
    return response


@login_required
def export_tshirt_stock_pdf(request):
    rows = TshirtStock.objects.select_related("brand").order_by("brand__name", "size").values_list("brand__name", "size", "available_quantity", "allocated_quantity", "low_stock_threshold")
    formatted = ((brand, size, available, allocated, threshold, "LOW" if available <= threshold else "OK") for brand, size, available, allocated, threshold in rows)
    return _pdf_response("tshirt_stock.pdf", "Next Toppers T-shirt Stock", ["Brand", "Size", "Available", "Allocated", "Threshold", "Status"], formatted, [220, 70, 90, 90, 90, 70])


@login_required
def export_low_stock_excel(request):
    rows = []
    for stock in TshirtStock.objects.select_related("brand").order_by("brand__name", "size"):
        if stock.is_low_stock:
            rows.append((stock.brand.name, stock.size, stock.available_quantity, stock.allocated_quantity, stock.low_stock_threshold))
    return _excel_response("low_stock.xlsx", "Low Stock", ["Brand", "Size", "Available", "Allocated", "Threshold"], rows)


@role_required(User.Role.SUPER_ADMIN)
def export_audit_excel(request):
    rows = AuditLog.objects.select_related("actor").values_list("created_at", "actor__employee_id", "action", "entity_type", "entity_id", "description")
    return _excel_response("audit_history.xlsx", "Audit History", ["Date and time", "Actor", "Action", "Entity type", "Entity ID", "Description"], rows)
