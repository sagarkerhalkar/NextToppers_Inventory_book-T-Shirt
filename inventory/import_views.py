from io import BytesIO

from django.contrib import messages
from django.http import Http404, HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .import_forms import BulkImportForm
from .import_services import run_import
from .models import User
from .permissions import role_required
from .services import audit


TEMPLATES = {
    "employees": {
        "filename": "employee_import_template.xlsx",
        "sheet": "Employees",
        "headers": [
            "employee_id", "full_name", "mobile_number", "default_tshirt_size",
            "email", "role", "department", "designation", "joining_date",
            "office_location", "temporary_password",
        ],
        "example": [
            "NXTTP0043", "Example Employee", "+919876543210", "L",
            "employee@example.com", "STAFF", "", "", "2026-07-22", "", "ChangeMe123!",
        ],
    },
    "books": {
        "filename": "book_import_template.xlsx",
        "sheet": "Books",
        "headers": [
            "asset_id", "book_name", "class_name", "stream_name", "isbn",
            "purchase_date", "bill_number", "condition",
        ],
        "example": ["", "Physics Part 1", "11", "Science", "9780000000000", "2026-07-22", "BILL-001", "GOOD"],
    },
    "tshirts": {
        "filename": "tshirt_stock_import_template.xlsx",
        "sheet": "Tshirt Stock",
        "headers": [
            "brand", "size", "quantity", "purchase_date", "vendor", "bill_number",
            "total_cost", "free_allowance", "low_stock_threshold",
        ],
        "example": ["Next Toppers", "L", 25, "2026-07-22", "Vendor Name", "TS-001", 7500, 5, 5],
    },
}


@role_required(User.Role.ADMIN, User.Role.SUPER_ADMIN)
def bulk_import(request):
    form = BulkImportForm(request.POST or None, request.FILES or None)
    result = None
    if request.method == "POST" and form.is_valid():
        try:
            result = run_import(
                form.cleaned_data["import_type"],
                form.cleaned_data["excel_file"],
                request.user,
            )
            audit(
                request.user,
                "BULK_IMPORT_COMPLETED",
                request.user,
                f"{result.import_type}: {result.successful} successful, {result.failed} failed",
                metadata={
                    "import_type": result.import_type,
                    "total_rows": result.total_rows,
                    "created": result.created,
                    "updated": result.updated,
                    "failed": result.failed,
                },
            )
            if result.failed:
                messages.warning(request, f"Import completed with {result.failed} row error(s).")
            else:
                messages.success(request, f"Import completed successfully. {result.successful} row(s) processed.")
        except Exception as exc:
            form.add_error("excel_file", str(exc))
    return render(request, "inventory/imports/index.html", {"form": form, "result": result})


@role_required(User.Role.ADMIN, User.Role.SUPER_ADMIN)
def download_import_template(request, import_type):
    try:
        definition = TEMPLATES[import_type]
    except KeyError as exc:
        raise Http404("Unknown import template") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = definition["sheet"]
    sheet.append(definition["headers"])
    sheet.append(definition["example"])
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 3
        sheet.column_dimensions[column[0].column_letter].width = min(max(width, 12), 35)

    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{definition["filename"]}"'
    return response
