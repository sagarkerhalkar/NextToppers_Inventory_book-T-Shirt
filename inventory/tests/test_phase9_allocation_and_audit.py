from datetime import date, datetime, timedelta
from io import BytesIO
from tempfile import TemporaryDirectory

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from inventory.forms import EmployeeRecordForm
from inventory.models import (
    AuditLog,
    Book,
    Employee,
    TshirtAllocation,
    TshirtBrand,
    TshirtPurchase,
    TshirtStock,
    User,
)
from inventory.services import allocate_book, free_entitlement, issue_free_tshirts


class PhaseNineAllocationAndAuditTests(TestCase):
    def setUp(self):
        self.media_dir = TemporaryDirectory()
        self.media_override = override_settings(MEDIA_ROOT=self.media_dir.name)
        self.media_override.enable()
        self.addCleanup(self.media_override.disable)
        self.addCleanup(self.media_dir.cleanup)

        self.admin = User.objects.create_user(
            employee_id="NXTTP0001",
            full_name="Admin User",
            mobile_number="+919876543210",
            password="1234",
            role=User.Role.ADMIN,
        )
        self.staff = User.objects.create_user(
            employee_id="NXTTP0002",
            full_name="Data Entry User",
            mobile_number="+919876543211",
            password="1234",
            role=User.Role.STAFF,
        )
        self.employee = Employee.objects.create(
            employee_id="NXTTP0100",
            full_name="Sagar Employee",
            mobile_number="+919876543212",
            default_tshirt_size="L",
        )
        self.brand = TshirtBrand.objects.create(
            name="Next Toppers",
            free_quantity_rolling_12_months=5,
        )
        self.stock = TshirtStock.objects.create(
            brand=self.brand,
            size="L",
            available_quantity=20,
        )
        self.book = Book.objects.create(
            asset_id="NTB-100",
            name="Physics",
            publication_name="NT Publication",
            subject="Physics",
            purchase_date=date(2025, 4, 10),
            bill_photo=SimpleUploadedFile("book_bill.jpg", b"book-bill", content_type="image/jpeg"),
            book_photo=SimpleUploadedFile("book_photo.jpg", b"book-photo", content_type="image/jpeg"),
            created_by=self.admin,
        )

    def test_employee_autocomplete_searches_large_employee_master(self):
        self.client.force_login(self.staff)
        response = self.client.get(reverse("inventory:employee_autocomplete"), {"q": "Sagar"})
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["results"][0]["id"], self.employee.pk)
        self.assertIn("NXTTP0100", payload["results"][0]["text"])

    def test_stock_autocomplete_displays_available_quantity(self):
        self.client.force_login(self.staff)
        response = self.client.get(reverse("inventory:tshirt_stock_autocomplete"), {"q": "Next"})
        self.assertEqual(response.status_code, 200)
        self.assertIn("Available 20", response.json()["results"][0]["text"])

    def test_entitlement_dates_are_optional_and_admin_editable(self):
        form = EmployeeRecordForm(
            {
                "employee_id": "NXTTP0101",
                "full_name": "No Fixed Period",
                "mobile_number": "",
                "default_tshirt_size": "M",
                "is_active": "on",
            },
            actor=self.admin,
        )
        self.assertTrue(form.is_valid(), form.errors)
        employee = form.save()
        self.assertIsNone(employee.tshirt_entitlement_start_date)
        self.assertIsNone(employee.tshirt_entitlement_end_date)

        form = EmployeeRecordForm(
            {
                "employee_id": employee.employee_id,
                "full_name": employee.full_name,
                "mobile_number": "",
                "default_tshirt_size": "M",
                "tshirt_entitlement_start_date": "2025-04-01",
                "tshirt_entitlement_end_date": "2026-03-31",
                "is_active": "on",
            },
            instance=employee,
            actor=self.admin,
        )
        self.assertTrue(form.is_valid(), form.errors)
        employee = form.save()
        self.assertEqual(employee.tshirt_entitlement_start_date, date(2025, 4, 1))
        self.assertEqual(employee.tshirt_entitlement_end_date, date(2026, 3, 31))

    def test_employee_fixed_entitlement_period_controls_free_issue(self):
        self.employee.tshirt_entitlement_start_date = date(2025, 4, 1)
        self.employee.tshirt_entitlement_end_date = date(2026, 3, 31)
        self.employee.save()
        inside = timezone.make_aware(datetime(2025, 7, 1, 10, 0))
        issue_free_tshirts(
            employee=self.employee,
            stock=self.stock,
            quantity=2,
            actor=self.staff,
            issued_at=inside,
        )
        summary = free_entitlement(self.employee, self.brand, as_of=inside)
        self.assertEqual(summary["used"], 2)
        self.assertEqual(summary["remaining"], 3)

        outside = timezone.make_aware(datetime(2026, 4, 1, 10, 0))
        with self.assertRaisesMessage(ValueError, "entitlement period"):
            issue_free_tshirts(
                employee=self.employee,
                stock=self.stock,
                quantity=1,
                actor=self.staff,
                issued_at=outside,
            )

    def test_past_book_allocation_is_recorded_and_audited(self):
        allocated_at = timezone.now() - timedelta(days=30)
        allocation = allocate_book(
            book=self.book,
            employee=self.employee,
            actor=self.staff,
            allocated_at=allocated_at,
        )
        self.assertAlmostEqual(allocation.allocated_at.timestamp(), allocated_at.timestamp(), delta=1)
        log = AuditLog.objects.get(action="BOOK_ALLOCATED")
        self.assertTrue(log.metadata["backdated"])

    def test_admin_can_correct_available_stock_with_reason(self):
        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("inventory:tshirt_stock_correct", args=[self.stock.pk]),
            {"available_quantity": 17, "correction_reason": "Physical count correction"},
        )
        self.assertEqual(response.status_code, 302)
        self.stock.refresh_from_db()
        self.assertEqual(self.stock.available_quantity, 17)
        log = AuditLog.objects.get(action="TSHIRT_AVAILABLE_STOCK_CORRECTED")
        self.assertEqual(log.metadata["old_available"], 20)
        self.assertEqual(log.metadata["new_available"], 17)

    def test_data_entry_user_cannot_correct_stock(self):
        self.client.force_login(self.staff)
        response = self.client.get(reverse("inventory:tshirt_stock_correct", args=[self.stock.pk]))
        self.assertIn(response.status_code, {302, 403})

    def _create_paid_request_and_purchase(self):
        paid = TshirtAllocation.objects.create(
            employee_record=self.employee,
            stock=self.stock,
            quantity=1,
            issue_type=TshirtAllocation.IssueType.PAID,
            status=TshirtAllocation.Status.PENDING,
            requested_by=self.staff,
            requested_at=timezone.make_aware(datetime(2025, 5, 5, 10, 30)),
            payment_amount="500.00",
            payment_date=date(2025, 5, 5),
            payment_proof=SimpleUploadedFile("payment.pdf", b"payment-proof", content_type="application/pdf"),
            hr_approval_proof=SimpleUploadedFile("hr.pdf", b"hr-proof", content_type="application/pdf"),
        )
        purchase = TshirtPurchase.objects.create(
            stock=self.stock,
            purchase_date=date(2025, 4, 2),
            vendor="Audit Vendor",
            bill_number="TS-BILL-1",
            bill_photo=SimpleUploadedFile("tshirt_bill.jpg", b"tshirt-bill", content_type="image/jpeg"),
            quantity=10,
            total_cost="2500.00",
            created_by=self.admin,
        )
        return paid, purchase

    def test_admin_can_view_and_download_paid_documents(self):
        paid, _purchase = self._create_paid_request_and_purchase()
        self.client.force_login(self.admin)
        view_response = self.client.get(reverse("inventory:paid_tshirt_document", args=[paid.pk, "payment"]))
        self.addCleanup(view_response.close)
        self.assertEqual(view_response.status_code, 200)
        self.assertIn("inline", view_response["Content-Disposition"])
        download_response = self.client.get(
            reverse("inventory:paid_tshirt_document", args=[paid.pk, "payment"]),
            {"download": "1"},
        )
        self.addCleanup(download_response.close)
        self.assertEqual(download_response.status_code, 200)
        self.assertIn("attachment", download_response["Content-Disposition"])

    def test_data_entry_user_cannot_open_protected_documents(self):
        paid, _purchase = self._create_paid_request_and_purchase()
        self.client.force_login(self.staff)
        response = self.client.get(reverse("inventory:paid_tshirt_document", args=[paid.pk, "payment"]))
        self.assertIn(response.status_code, {302, 403})

    def test_employee_old_calendar_year_report_filters_correctly(self):
        allocate_book(
            book=self.book,
            employee=self.employee,
            actor=self.staff,
            allocated_at=timezone.make_aware(datetime(2025, 6, 1, 9, 0)),
        )
        self.client.force_login(self.admin)
        response = self.client.get(
            reverse("inventory:download_activity_report"),
            {
                "employee": self.employee.pk,
                "report_type": "book",
                "period": "calendar_year",
                "year": "2025",
                "format": "xlsx",
            },
        )
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        self.addCleanup(workbook.close)
        self.assertEqual(workbook["Summary"]["B2"].value, "NXTTP0100 — Sagar Employee")
        self.assertEqual(workbook["Book Activity"].max_row, 2)

    def test_audit_evidence_register_and_excel_contain_document_links(self):
        paid, purchase = self._create_paid_request_and_purchase()
        self.client.force_login(self.admin)
        register = self.client.get(reverse("inventory:audit_evidence_register"))
        self.assertEqual(register.status_code, 200)
        self.assertContains(register, "Book Bill")
        self.assertContains(register, "T-shirt Purchase Bill")
        self.assertContains(register, "Paid T-shirt Payment Proof")

        export = self.client.get(reverse("inventory:audit_evidence_export"), {"format": "xlsx"})
        self.assertEqual(export.status_code, 200)
        workbook = load_workbook(BytesIO(export.content))
        self.addCleanup(workbook.close)
        sheet = workbook["Audit Evidence"]
        self.assertGreaterEqual(sheet.max_row, 6)
        self.assertTrue(any(sheet.cell(row, 6).hyperlink for row in range(2, sheet.max_row + 1)))
        self.assertTrue(any(sheet.cell(row, 7).hyperlink for row in range(2, sheet.max_row + 1)))

        bill = self.client.get(reverse("inventory:tshirt_purchase_document", args=[purchase.pk]), {"download": "1"})
        self.addCleanup(bill.close)
        self.assertEqual(bill.status_code, 200)
        self.assertIn("attachment", bill["Content-Disposition"])
