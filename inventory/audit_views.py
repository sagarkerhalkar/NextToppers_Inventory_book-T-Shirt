from datetime import datetime
from io import BytesIO

from django.core.paginator import Paginator
from django.db import models
from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import Book, Employee, TshirtAllocation, TshirtPurchase, User
from .permissions import role_required


DOCUMENT_TYPES = {
    "book_bill": "Book Bill",
    "book_photo": "Book Photo",
    "tshirt_purchase_bill": "T-shirt Purchase Bill",
    "payment_proof": "Paid T-shirt Payment Proof",
    "hr_approval": "Paid T-shirt HR Approval",
}


def _parse_date(value, label):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(f"{label} must use YYYY-MM-DD format.") from exc


def _date_in_range(value, start_date, end_date):
    if value is None:
        return not start_date and not end_date
    if hasattr(value, "date") and not isinstance(value, datetime):
        normalized = value
    elif isinstance(value, datetime):
        normalized = timezone.localtime(value).date() if timezone.is_aware(value) else value.date()
    else:
        normalized = value
    if start_date and normalized < start_date:
        return False
    if end_date and normalized > end_date:
        return False
    return True


def _match_query(query, *values):
    if not query:
        return True
    needle = query.lower()
    return any(needle in str(value or "").lower() for value in values)


def _document_urls(request, view_name, args):
    base = request.build_absolute_uri(reverse(view_name, args=args))
    return base, f"{base}?download=1"


def _evidence_rows(request):
    query = request.GET.get("q", "").strip()
    document_type = request.GET.get("document_type", "all")
    employee_value = request.GET.get("employee", "").strip()
    start_date = _parse_date(request.GET.get("start_date", ""), "Start date")
    end_date = _parse_date(request.GET.get("end_date", ""), "End date")
    if start_date and end_date and start_date > end_date:
        raise ValueError("Start date cannot be after end date.")
    employee = None
    if employee_value:
        try:
            employee = get_object_or_404(Employee, pk=int(employee_value))
        except (TypeError, ValueError):
            raise ValueError("Invalid employee selection.")

    rows = []
    allowed = set(DOCUMENT_TYPES) if document_type == "all" else {document_type}
    if not allowed.issubset(DOCUMENT_TYPES):
        raise ValueError("Unknown document type.")

    if "book_bill" in allowed or "book_photo" in allowed:
        books = Book.objects.filter(is_active=True).only(
            "pk", "asset_id", "name", "publication_name", "subject", "purchase_date", "created_at", "bill_number", "bill_photo", "book_photo"
        )
        for book in books.iterator(chunk_size=500):
            event_date = book.purchase_date or timezone.localtime(book.created_at).date()
            if not _date_in_range(event_date, start_date, end_date):
                continue
            if not _match_query(query, book.asset_id, book.name, book.publication_name, book.subject, book.bill_number):
                continue
            if "book_bill" in allowed and book.bill_photo:
                view_url, download_url = _document_urls(request, "inventory:book_document", [book.pk, "bill"])
                rows.append({
                    "date": event_date,
                    "type": "Book Bill",
                    "reference": book.asset_id,
                    "description": f"{book.name} · Bill {book.bill_number or '-'}",
                    "employee": "",
                    "view_url": view_url,
                    "download_url": download_url,
                })
            if "book_photo" in allowed and book.book_photo:
                view_url, download_url = _document_urls(request, "inventory:book_document", [book.pk, "photo"])
                rows.append({
                    "date": event_date,
                    "type": "Book Photo",
                    "reference": book.asset_id,
                    "description": book.name,
                    "employee": "",
                    "view_url": view_url,
                    "download_url": download_url,
                })

    if "tshirt_purchase_bill" in allowed:
        purchases = TshirtPurchase.objects.select_related("stock", "stock__brand").filter(bill_photo__isnull=False).exclude(bill_photo="")
        for purchase in purchases.iterator(chunk_size=500):
            event_date = purchase.purchase_date
            if not _date_in_range(event_date, start_date, end_date):
                continue
            if not _match_query(query, purchase.stock.brand.name, purchase.stock.size, purchase.vendor, purchase.bill_number):
                continue
            view_url, download_url = _document_urls(request, "inventory:tshirt_purchase_document", [purchase.pk])
            rows.append({
                "date": event_date,
                "type": "T-shirt Purchase Bill",
                "reference": f"PUR-{purchase.pk}",
                "description": f"{purchase.stock.brand.name} · Size {purchase.stock.size} · Vendor {purchase.vendor or '-'} · Bill {purchase.bill_number or '-'}",
                "employee": "",
                "view_url": view_url,
                "download_url": download_url,
            })

    if "payment_proof" in allowed or "hr_approval" in allowed:
        allocations = TshirtAllocation.objects.select_related("employee_record", "employee", "stock", "stock__brand").filter(issue_type=TshirtAllocation.IssueType.PAID)
        if employee is not None:
            allocations = allocations.filter(employee_record=employee)
        for allocation in allocations.iterator(chunk_size=500):
            recipient = allocation.recipient
            employee_text = f"{recipient.employee_id} · {recipient.full_name}" if recipient else ""
            event_date = allocation.payment_date or timezone.localtime(allocation.requested_at).date()
            if not _date_in_range(event_date, start_date, end_date):
                continue
            if not _match_query(
                query,
                employee_text,
                allocation.stock.brand.name,
                allocation.stock.size,
                allocation.payment_amount,
                allocation.status,
            ):
                continue
            if "payment_proof" in allowed and allocation.payment_proof:
                view_url, download_url = _document_urls(request, "inventory:paid_tshirt_document", [allocation.pk, "payment"])
                rows.append({
                    "date": event_date,
                    "type": "Paid T-shirt Payment Proof",
                    "reference": f"PAID-{allocation.pk}",
                    "description": f"{allocation.stock.brand.name} · Size {allocation.stock.size} · Amount ₹{allocation.payment_amount or 0}",
                    "employee": employee_text,
                    "view_url": view_url,
                    "download_url": download_url,
                })
            if "hr_approval" in allowed and allocation.hr_approval_proof:
                view_url, download_url = _document_urls(request, "inventory:paid_tshirt_document", [allocation.pk, "hr"])
                rows.append({
                    "date": event_date,
                    "type": "Paid T-shirt HR Approval",
                    "reference": f"PAID-{allocation.pk}",
                    "description": f"{allocation.stock.brand.name} · Size {allocation.stock.size} · {allocation.get_status_display()}",
                    "employee": employee_text,
                    "view_url": view_url,
                    "download_url": download_url,
                })

    rows.sort(key=lambda row: (row["date"], row["type"], row["reference"]), reverse=True)
    return rows, employee, start_date, end_date, document_type, query


@role_required(User.Role.ADMIN, User.Role.SUPER_ADMIN)
def audit_evidence_register(request):
    try:
        rows, employee, start_date, end_date, document_type, query = _evidence_rows(request)
    except ValueError as exc:
        return HttpResponseBadRequest(str(exc))
    page_size = request.GET.get("page_size", "20")
    page_size = 30 if page_size == "30" else 20
    paginator = Paginator(rows, page_size)
    page_obj = paginator.get_page(request.GET.get("page"))
    preserved = request.GET.copy()
    preserved.pop("page", None)
    return render(request, "inventory/reports/audit_evidence.html", {
        "rows": page_obj,
        "page_obj": page_obj,
        "paginator": paginator,
        "page_size": page_size,
        "page_query": preserved.urlencode(),
        "document_types": DOCUMENT_TYPES,
        "selected_type": document_type,
        "query": query,
        "employee": employee,
        "start_date": start_date,
        "end_date": end_date,
        "total_documents": len(rows),
    })


@role_required(User.Role.ADMIN, User.Role.SUPER_ADMIN)
def audit_evidence_export(request):
    output_format = request.GET.get("format", "xlsx")
    if output_format not in {"xlsx", "pdf"}:
        return HttpResponseBadRequest("Unknown export format.")
    try:
        rows, employee, start_date, end_date, _document_type, _query = _evidence_rows(request)
    except ValueError as exc:
        return HttpResponseBadRequest(str(exc))
    employee_slug = employee.employee_id if employee else "all"
    period_slug = f"{start_date or 'all'}_to_{end_date or 'all'}"

    if output_format == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Audit Evidence"
        sheet.append(["Date", "Document Type", "Reference", "Description", "Employee", "View Link", "Download Link"])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F766E")
        for row in rows:
            sheet.append([
                row["date"].strftime("%d-%m-%Y"), row["type"], row["reference"], row["description"], row["employee"],
                "View document", "Download document",
            ])
            current = sheet.max_row
            sheet.cell(current, 6).hyperlink = row["view_url"]
            sheet.cell(current, 6).style = "Hyperlink"
            sheet.cell(current, 7).hyperlink = row["download_url"]
            sheet.cell(current, 7).style = "Hyperlink"
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        widths = [14, 30, 18, 60, 32, 20, 22]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
        output = BytesIO()
        workbook.save(output)
        response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="audit_evidence_{employee_slug}_{period_slug}.xlsx"'
        return response

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="audit_evidence_{employee_slug}_{period_slug}.pdf"'
    document = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=10*mm, leftMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    styles = getSampleStyleSheet()
    normal = styles["BodyText"]
    normal.fontSize = 6.2
    story = [
        Paragraph("Next Toppers Finance / ISO Audit Evidence Register", styles["Title"]),
        Paragraph(f"Employee: {employee.employee_id + ' · ' + employee.full_name if employee else 'All employees'}", styles["Normal"]),
        Paragraph(f"Documents: {len(rows)}", styles["Normal"]),
        Spacer(1, 5*mm),
    ]
    data = [["Date", "Type", "Reference", "Description", "Employee", "View", "Download"]]
    for row in rows:
        data.append([
            row["date"].strftime("%d-%m-%Y"),
            row["type"],
            row["reference"],
            Paragraph(row["description"], normal),
            Paragraph(row["employee"], normal),
            Paragraph(f'<link href="{row["view_url"]}">Open</link>', normal),
            Paragraph(f'<link href="{row["download_url"]}">Download</link>', normal),
        ])
    table = Table(data, repeatRows=1, colWidths=[22*mm, 35*mm, 25*mm, 70*mm, 48*mm, 22*mm, 26*mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 6.2),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#B0BEC5")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#E8F5F3")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(table)
    document.build(story)
    return response
