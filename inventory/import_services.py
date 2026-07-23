from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.utils import timezone
from openpyxl import load_workbook

from .models import Book, Employee, TshirtBrand, TshirtPurchase, TshirtStock, User
from .services import audit


@dataclass
class ImportResult:
    import_type: str
    total_rows: int = 0
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[dict] = field(default_factory=list)

    @property
    def failed(self):
        return len(self.errors)

    @property
    def successful(self):
        return self.created + self.updated

    def add_error(self, row_number, message):
        self.errors.append({"row": row_number, "message": str(message)})


def _normalise_header(value):
    return str(value or "").strip().lower().replace(" ", "_").replace("-", "_")


def _clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _date_value(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Invalid date '{value}'. Use YYYY-MM-DD or DD-MM-YYYY.")


def _integer(value, field_name, minimum=0):
    try:
        number = int(value)
    except (TypeError, ValueError):
        raise ValueError(f"{field_name} must be a whole number.")
    if number < minimum:
        raise ValueError(f"{field_name} must be at least {minimum}.")
    return number


def _decimal(value, field_name):
    if value in (None, ""):
        return Decimal("0.00")
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        raise ValueError(f"{field_name} must be a valid amount.")


def _rows(uploaded_file):
    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not raw_headers:
        raise ValueError("The Excel file is empty.")
    headers = [_normalise_header(value) for value in raw_headers]
    if not any(headers):
        raise ValueError("The first row must contain column headings.")
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        record = {headers[index]: _clean(values[index]) if index < len(values) else "" for index in range(len(headers)) if headers[index]}
        if any(value not in (None, "") for value in record.values()):
            yield row_number, record


def _require_columns(record, required):
    missing = [name for name in required if record.get(name, "") in (None, "")]
    if missing:
        raise ValueError("Missing required value(s): " + ", ".join(missing))


def import_employees(uploaded_file, actor):
    result = ImportResult("employees")
    valid_sizes = {choice for choice, _ in User.TshirtSize.choices}
    for row_number, record in _rows(uploaded_file):
        result.total_rows += 1
        try:
            _require_columns(record, ["employee_id", "full_name", "default_tshirt_size"])
            employee_id = str(record["employee_id"]).upper()
            mobile = str(record.get("mobile_number") or "").strip() or None
            size = str(record["default_tshirt_size"]).upper()
            if size not in valid_sizes:
                raise ValueError(f"Invalid T-shirt size '{size}'.")
            if Employee.objects.filter(employee_id=employee_id).exists():
                raise ValueError(f"Employee ID {employee_id} already exists.")
            if mobile and Employee.objects.filter(mobile_number=mobile).exists():
                raise ValueError(f"Mobile number {mobile} already exists.")
            with transaction.atomic():
                employee = Employee(
                    employee_id=employee_id,
                    full_name=str(record["full_name"]),
                    mobile_number=mobile,
                    email=str(record.get("email") or ""),
                    department=str(record.get("department") or ""),
                    designation=str(record.get("designation") or ""),
                    joining_date=_date_value(record.get("joining_date")),
                    office_location=str(record.get("office_location") or ""),
                    default_tshirt_size=size,
                    notes=str(record.get("notes") or ""),
                    is_active=True,
                )
                employee.save()
                audit(actor, "EMPLOYEE_IMPORTED", employee, f"Imported non-login employee {employee.employee_id} from Excel")
            result.created += 1
        except (ValueError, ValidationError, IntegrityError) as exc:
            result.add_error(row_number, exc)
    return result


def import_books(uploaded_file, actor):
    result = ImportResult("books")
    valid_conditions = {choice for choice, _ in Book.Condition.choices}
    for row_number, record in _rows(uploaded_file):
        result.total_rows += 1
        try:
            _require_columns(record, ["book_name"])
            asset_id = str(record.get("asset_id") or "").upper()
            if asset_id and Book.objects.filter(asset_id=asset_id).exists():
                raise ValueError(f"Book Asset ID {asset_id} already exists.")
            condition = str(record.get("condition") or Book.Condition.GOOD).upper()
            if condition not in valid_conditions:
                raise ValueError(f"Invalid Book condition '{condition}'.")
            with transaction.atomic():
                book = Book(
                    asset_id=asset_id,
                    name=str(record["book_name"]),
                    publication_name=str(record.get("publication_name") or ""),
                    subject=str(record.get("subject") or ""),
                    class_name=str(record.get("class_name") or ""),
                    stream_name=str(record.get("stream_name") or ""),
                    isbn=str(record.get("isbn") or ""),
                    purchase_date=_date_value(record.get("purchase_date")),
                    bill_number=str(record.get("bill_number") or ""),
                    condition=condition,
                    created_by=actor,
                )
                book.full_clean(exclude=["asset_id"] if not asset_id else None)
                book.save()
                audit(actor, "BOOK_IMPORTED", book, f"Imported {book.asset_id} from Excel")
            result.created += 1
        except (ValueError, ValidationError, IntegrityError) as exc:
            result.add_error(row_number, exc)
    return result


def import_tshirt_stock(uploaded_file, actor):
    result = ImportResult("tshirts")
    valid_sizes = {choice for choice, _ in User.TshirtSize.choices}
    for row_number, record in _rows(uploaded_file):
        result.total_rows += 1
        try:
            _require_columns(record, ["brand", "size", "quantity"])
            brand_name = str(record["brand"]).strip()
            size = str(record["size"]).upper()
            if size not in valid_sizes:
                raise ValueError(f"Invalid T-shirt size '{size}'.")
            quantity = _integer(record["quantity"], "Quantity", minimum=1)
            free_allowance = _integer(record.get("free_allowance") or 0, "Free allowance", minimum=0)
            threshold = _integer(record.get("low_stock_threshold") or 5, "Low-stock threshold", minimum=0)
            with transaction.atomic():
                brand = TshirtBrand.objects.filter(name__iexact=brand_name).first()
                brand_created = brand is None
                if brand_created:
                    brand = TshirtBrand.objects.create(name=brand_name, free_quantity_rolling_12_months=free_allowance)
                elif record.get("free_allowance") not in (None, ""):
                    brand.free_quantity_rolling_12_months = free_allowance
                    brand.save(update_fields=["free_quantity_rolling_12_months", "updated_at"])
                stock, stock_created = TshirtStock.objects.get_or_create(brand=brand, size=size, defaults={"low_stock_threshold": threshold})
                if not stock_created and record.get("low_stock_threshold") not in (None, ""):
                    stock.low_stock_threshold = threshold
                stock.available_quantity += quantity
                stock.save()
                purchase = TshirtPurchase.objects.create(stock=stock, purchase_date=_date_value(record.get("purchase_date")) or timezone.localdate(), vendor=str(record.get("vendor") or ""), bill_number=str(record.get("bill_number") or ""), quantity=quantity, total_cost=_decimal(record.get("total_cost"), "Total cost"), created_by=actor)
                audit(actor, "TSHIRT_STOCK_IMPORTED", purchase, f"Imported {quantity} units for {stock}")
            result.created += 1 if stock_created else 0
            result.updated += 0 if stock_created else 1
        except (ValueError, ValidationError, IntegrityError) as exc:
            result.add_error(row_number, exc)
    return result


def run_import(import_type, uploaded_file, actor):
    handlers = {"employees": import_employees, "books": import_books, "tshirts": import_tshirt_stock}
    try:
        handler = handlers[import_type]
    except KeyError as exc:
        raise ValueError("Unsupported import type.") from exc
    return handler(uploaded_file, actor)
